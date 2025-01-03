import re
import uuid
from typing import List, Dict

import pandas
from fastapi import Depends, HTTPException, Request, UploadFile
import os
import http
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.workbook import Workbook

from src.domains.calculations.calculation_interface import (
    ICalculationRepository,
    ICalculationUseCase,
)
from src.domains.calculations.calculation_repository import CalculationRepository
from src.domains.calculations.entities.va_slot_calculation_details import (
    SlotCalculationDetail,
)
from src.domains.calculations.entities.va_slot_calculations import SlotCalculation
from src.domains.masters.entities.va_models import Model
from src.domains.masters.master_interface import IMasterRepository
from src.domains.masters.master_repository import MasterRepository
from src.models.requests.calculation_request import (
    GetCalculationRequest,
    UpdateCalculationRequest,
)
from src.models.responses.basic_response import TextValueResponse
from src.models.responses.calculation_response import (
    GetCalculationDetailResponse,
    GetCalculationResponse,
    GetCalculationStockPilotResponse,
    GetCalculationDetailMonthsResponse,
)
from src.models.responses.master_response import (
    CategoryResponse,
    ModelResponse,
    SegmentResponse,
)
from src.shared.enums import Database
from src.shared.utils.database_utils import begin_transaction, commit
from src.shared.utils.date import get_month_difference, is_date_string_format
from src.shared.utils.excel import (
    get_header_column_index,
    get_worksheet,
    open_excel_workbook,
)
from src.shared.utils.file_utils import (
    clear_directory,
    get_file_extension,
    save_upload_file,
)
from src.shared.utils.storage_utils import is_file_exist, get_full_path
from src.shared.utils.xid import generate_xid
from pathlib import Path


class CalculationUseCase(ICalculationUseCase):

    def __init__(
        self,
        calculation_repo: ICalculationRepository = Depends(CalculationRepository),
        master_repository: IMasterRepository = Depends(MasterRepository),
    ):
        self.calculation_repo = calculation_repo
        self.master_repository = master_repository

    def upsert_bo_soa_oc_booking_prospect(
        self, request, file: str, month: int, year: int
    ):
        if not is_file_exist(file):
            raise HTTPException(
                status_code=http.HTTPStatus.INTERNAL_SERVER_ERROR,
                detail="Excel file is not found",
            )

        df = pandas.read_excel(os.getcwd() + "/storage" + file)
        columns = [
            "SO Number",
            "Status SO",
            "Status Pilot",
            "Dealer",
            "Model",
            "Cat",
            "Region",
            "Customer Name",
            "Vin Year Rev",
            "SRC",
            "Source",
            "Year",
        ]

        not_found_index = [i for i in columns if i not in df.columns.tolist()]
        forecast_months = [
            i for i in df.columns.tolist() if re.match(r"^\d{4}-(0[1-9]|1[0-2])$", i)
        ]

        if len(not_found_index) > 0:
            raise HTTPException(
                http.HTTPStatus.BAD_REQUEST,
                detail=f"{not_found_index[0]} field is required.",
            )

        begin_transaction(request, Database.VEHICLE_ALLOCATION)

        calculation_details: List[SlotCalculationDetail] = []
        calculation_detail_map = {}

        model_dict: Dict[str, Model] = {}

        for idx, row in df.iterrows():
            so_number = row["SO Number"]
            status_so = row["Status SO"]
            status_pilot = row["Status Pilot"]
            dealer = row["Dealer"]
            model_id = row["Model"]
            cat = row["Cat"]
            region = row["Region"]
            customer_name = row["Customer Name"]
            vin_year_rev = row["Vin Year Rev"]
            src = row["SRC"]
            source = row["Source"]
            row_year = row["Year"]

            if model_id not in model_dict:
                model = self.master_repository.find_model(request, model_id)
                if model is None:
                    raise HTTPException(
                        http.HTTPStatus.BAD_REQUEST,
                        detail=f"Model {model_id} is not found",
                    )
                model_dict[model_id] = model

            model = model_dict[model_id]

            is_soa = (
                so_number[:2].upper() == "SO"
                and status_so.upper() == "ACCEPTED"
                and status_pilot.upper() == "PILOT"
                and source.upper() == "FCST ORDER"
            )

            is_so = (
                so_number[:2].upper() == "SO"
                and status_so.upper() == "PROSPECT"
                and status_pilot.upper() == "PILOT"
                and source.upper() == "FCST ORDER"
            )

            is_oc = (
                so_number[:2].upper() != "SO"
                and (status_so.upper() == "ACCEPTED" or status_so.upper() == "PROSPECT")
                and status_pilot.upper() == "PILOT"
                and source.upper() == "FCST ORDER"
            )

            is_booking = (
                so_number[:2].upper() != "SO" and source.upper() == "URGENT ORDER"
            )

            for j in forecast_months:
                forecast_month = get_month_difference(f"{year}-{month}", j)

                if forecast_month < 0:
                    raise HTTPException(
                        status_code=http.HTTPStatus.BAD_REQUEST,
                        detail=f"Forecast month cannot be less than the current month",
                    )

                if model.id not in calculation_detail_map:
                    calculation_detail_map[model.id] = {}

                if forecast_month not in calculation_detail_map[model.id]:
                    calculation_detail_map[model.id][forecast_month] = (
                        SlotCalculationDetail(
                            model_id=model.id,
                            forecast_month=forecast_month,
                            soa=0,
                            bo=0,
                            oc=0,
                            so=0,
                            booking_prospect=0,
                        )
                    )

                calculation_detail_map[model.id][forecast_month].soa += (
                    row[j] if is_soa and not pandas.isna(row[j]) else 0
                )
                calculation_detail_map[model.id][forecast_month].bo += 0
                calculation_detail_map[model.id][forecast_month].oc += (
                    row[j] if is_oc and not pandas.isna(row[j]) else 0
                )
                calculation_detail_map[model.id][forecast_month].so += (
                    row[j] if is_so and not pandas.isna(row[j]) else 0
                )
                calculation_detail_map[model.id][forecast_month].booking_prospect += (
                    row[j] if is_booking and not pandas.isna(row[j]) else 0
                )

        calculation = self.calculation_repo.find_calculation(
            request, month=month, year=year
        )

        if calculation is None:
            self.calculation_repo.create_calculation(
                request,
                SlotCalculation(month=month, year=year, details=calculation_details),
            )
        else:
            detail_maps = {}
            for i in calculation.details:
                if i.deletable == 0:
                    if i.model_id not in detail_maps:
                        detail_maps[i.model_id] = {}
                    if i.forecast_month not in detail_maps[i.model_id]:
                        detail_maps[i.model_id][i.forecast_month] = i

            for model_id, forecast_month_map in calculation_detail_map.items():
                if model_id in detail_maps:
                    for (
                        forecast_month,
                        calculation_detail,
                    ) in forecast_month_map.items():
                        calculation_detail.slot_calculation_id = calculation.id
                        if forecast_month not in detail_maps[model_id]:
                            self.calculation_repo.create_calculation_detail(
                                request, calculation_detail
                            )
                        else:
                            detail_maps[model_id][forecast_month].model_id = (
                                calculation_detail_map[model_id][
                                    forecast_month
                                ].model_id
                            )
                            detail_maps[model_id][forecast_month].soa = (
                                calculation_detail_map[model_id][forecast_month].soa
                            )
                            detail_maps[model_id][forecast_month].bo = (
                                calculation_detail_map[model_id][forecast_month].bo
                            )
                            detail_maps[model_id][forecast_month].oc = (
                                calculation_detail_map[model_id][forecast_month].oc
                            )
                            detail_maps[model_id][forecast_month].so = (
                                calculation_detail_map[model_id][forecast_month].so
                            )
                            detail_maps[model_id][forecast_month].booking_prospect = (
                                calculation_detail_map[model_id][
                                    forecast_month
                                ].booking_prospect
                            )
                else:
                    for (
                        forecast_month,
                        calculation_detail,
                    ) in forecast_month_map.items():
                        calculation_detail.slot_calculation_id = calculation.id
                        self.calculation_repo.create_calculation_detail(
                            request, calculation_detail
                        )

        commit(request, Database.VEHICLE_ALLOCATION)

    def upsert_take_off_data(
        self, request: Request, file: str, month: int, year: int
    ) -> None:
        if not is_file_exist(file):
            raise HTTPException(
                status_code=http.HTTPStatus.INTERNAL_SERVER_ERROR,
                detail="Excel file is not found",
            )

        df = pandas.read_excel(os.getcwd() + "/storage" + file)
        columns = [
            "Category",
            "Sub Name",
            "HMMI",
            "HMSI",
            "Class",
            "Sales Name",
            "Lot No",
            "Suffix/KIT",
            "Item name",
        ]

        not_found_index = [i for i in columns if i not in df.columns.tolist()]
        forecast_months = [
            i for i in df.columns.tolist() if re.match(r"^\d{4}-(0[1-9]|1[0-2])$", i)
        ]

        if len(not_found_index) > 0:
            raise HTTPException(
                http.HTTPStatus.BAD_REQUEST,
                detail=f"{not_found_index[0]} field is required.",
            )

        begin_transaction(request, Database.VEHICLE_ALLOCATION)

        calculation_details: List[SlotCalculationDetail] = []
        calculation_detail_map = {}

        model_dict: Dict[str, Model] = {}

        for idx, row in df.iterrows():
            model_id = row["Sales Name"]

            if model_id not in model_dict:
                model = self.master_repository.find_model(request, model_id)
                if model is None:
                    raise HTTPException(
                        http.HTTPStatus.BAD_REQUEST,
                        detail=f"Model {model_id} is not found",
                    )
                model_dict[model_id] = model

            model = model_dict[model_id]

            for j in forecast_months:
                forecast_month = get_month_difference(f"{year}-{month}", j)

                if forecast_month < 0:
                    raise HTTPException(
                        status_code=http.HTTPStatus.BAD_REQUEST,
                        detail=f"Forecast month cannot be less than the current month",
                    )

                if model.id not in calculation_detail_map:
                    calculation_detail_map[model.id] = {}

                if forecast_month not in calculation_detail_map[model.id]:
                    calculation_detail_map[model.id][forecast_month] = (
                        SlotCalculationDetail(
                            model_id=model.id,
                            forecast_month=forecast_month,
                            take_off=0,
                        )
                    )

                calculation_detail_map[model.id][forecast_month].take_off += (
                    row[j] if not pandas.isna(row[j]) else 0
                )

        calculation = self.calculation_repo.find_calculation(
            request, month=month, year=year
        )

        if calculation is None:
            self.calculation_repo.create_calculation(
                request,
                SlotCalculation(month=month, year=year, details=calculation_details),
            )
        else:
            detail_maps = {}
            for i in calculation.details:
                if i.deletable == 0:
                    if i.model_id not in detail_maps:
                        detail_maps[i.model_id] = {}
                    if i.forecast_month not in detail_maps[i.model_id]:
                        detail_maps[i.model_id][i.forecast_month] = i

            for model_id, forecast_month_map in calculation_detail_map.items():
                if model_id in detail_maps:
                    for (
                        forecast_month,
                        calculation_detail,
                    ) in forecast_month_map.items():
                        calculation_detail.slot_calculation_id = calculation.id
                        if forecast_month not in detail_maps[model_id]:
                            self.calculation_repo.create_calculation_detail(
                                request, calculation_detail
                            )
                        else:
                            detail_maps[model_id][forecast_month].model_id = (
                                calculation_detail_map[model_id][
                                    forecast_month
                                ].model_id
                            )
                            detail_maps[model_id][forecast_month].take_off = (
                                calculation_detail_map[model_id][
                                    forecast_month
                                ].take_off
                            )
                else:
                    for (
                        forecast_month,
                        calculation_detail,
                    ) in forecast_month_map.items():
                        calculation_detail.slot_calculation_id = calculation.id
                        self.calculation_repo.create_calculation_detail(
                            request, calculation_detail
                        )

        commit(request, Database.VEHICLE_ALLOCATION)

    def get_calculation_detail(
        self, request, get_calculation_request: GetCalculationRequest
    ) -> GetCalculationResponse | None:
        slot_calculation = self.calculation_repo.find_calculation(
            request,
            month=get_calculation_request.month,
            year=get_calculation_request.year,
        )

        if slot_calculation is None:
            raise HTTPException(
                status_code=http.HTTPStatus.NOT_FOUND, detail="Calculation not found"
            )

        model_map = {}

        for i in slot_calculation.details:
            if i.model_id not in model_map:
                model_map[i.model_id] = GetCalculationDetailMonthsResponse(
                    model_id=i.model_id,
                    months=[],
                    segment=TextValueResponse(
                        text=i.model.segment_id, value=i.model.segment_id
                    ),
                    category=TextValueResponse(
                        text=i.model.category_id, value=i.model.category_id
                    ),
                )

            model_map[i.model_id].months.append(
                GetCalculationDetailResponse(
                    calculation_slot_id=i.slot_calculation_id,
                    month=i.forecast_month,
                    take_off=0 if i.take_off is None else i.take_off,
                    bo=0 if i.bo is None else i.bo,
                    soa=0 if i.soa is None else i.soa,
                    oc=0 if i.oc is None else i.oc,
                    so=0 if i.so is None else i.so,
                    booking_prospect=(
                        0 if i.booking_prospect is None else i.booking_prospect
                    ),
                    slot_1=i.slot_1,
                    slot_2=i.slot_2,
                )
            )

        models = []

        for k, v in model_map.items():
            models.append(v)

        models = sorted(models, key=lambda x: (x.category.text, x.model_id))

        return GetCalculationResponse(models=models)

    def update_calculation_detail(
        self, request: Request, update_calculation_request: UpdateCalculationRequest
    ):
        begin_transaction(request, Database.VEHICLE_ALLOCATION)

        calculation_detail = self.calculation_repo.find_calculation_detail(
            request,
            update_calculation_request.slot_calculation_id,
            update_calculation_request.model_id,
            update_calculation_request.forecast_month,
        )

        if calculation_detail is None:
            raise HTTPException(
                status_code=http.HTTPStatus.NOT_FOUND,
                detail="Calculation detail not found",
            )

        calculation_detail.slot_1 = update_calculation_request.slot_1
        calculation_detail.slot_2 = update_calculation_request.slot_2

        commit(request, Database.VEHICLE_ALLOCATION)

    def download_booking_excel_template(
        self, request: Request, month: int, year: int
    ) -> str:
        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Template-Rundown"

        headers = [
            "SO Number",
            "Status SO",
            "Status Pilot",
            "Dealer",
            "Model",
            "Cat",
            "Region",
            "Customer Name",
            "Vin Year Rev",
            "SRC",
            "Source",
            "Year",
            "VRF QTY",
        ]

        for i in range(5):
            if i > 0:
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            headers.append(f"{year}-{month:02}")

        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[cell.column_letter].width = len(header) + 2

        relative_path = (
            "/temp/template/calculations/booking-" + str(uuid.uuid4()) + ".xlsx"
        )
        dest = get_full_path(relative_path)
        upload_dir = os.path.join(os.getcwd(), "storage/temp/template/calculations")
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        workbook.save(dest)
        return dest

    def download_takeoff_excel_template(self, request: Request, month: int, year: int):
        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Template-Rundown"

        headers = [
            "Category",
            "Sub Name",
            "HMMI",
            "HMSI",
            "Class",
            "Sales Name",
            "Lot No",
            "Suffix/KIT",
            "Item name",
        ]

        for i in range(7):
            if i > 0:
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            headers.append(f"{year}-{month:02}")

        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[cell.column_letter].width = len(header) + 2

        relative_path = (
            "/temp/template/calculations/takeoff-" + str(uuid.uuid4()) + ".xlsx"
        )
        dest = get_full_path(relative_path)
        upload_dir = os.path.join(os.getcwd(), "storage/temp/template/calculations")
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        workbook.save(dest)
        return dest
