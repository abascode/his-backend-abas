from typing import List
from fastapi import Depends, Request
import openpyxl
from src.domains.calculation.calculation_interface import ICalculationUseCase
from src.domains.forecast.forecast_interface import IForecastRepository
from src.domains.forecast.forecast_repository import ForecastRepository
from src.domains.master.entities.va_model import Model
from src.domains.master.master_interface import IMasterRepository
from src.domains.master.master_repository import MasterRepository
from src.models.requests.calculation_request import CalculationTemplateRequest
from openpyxl.styles import Alignment
import os

from src.shared.utils.date import convert_number_to_month


class CalculationUseCase(ICalculationUseCase):
    def __init__(
        self,
        forecast_repo: IForecastRepository = Depends(ForecastRepository),
        master_repo: IMasterRepository = Depends(MasterRepository),
    ):
        self.forecast_repo = forecast_repo
        self.master_repo = master_repo
        
    def generate_excel_month_header(self, requested_config: CalculationTemplateRequest):
        month_headers = []
    
        for i in range(len(requested_config.forecast)):
            month_header = '{} {} (N{})'.format(
                convert_number_to_month(requested_config.forecast[i]), 
                requested_config.year, 
                requested_config.forecast[i]
            )
            month_headers.append(month_header)
            
        return month_headers
    
    def generate_excel_business_header(self, business_header_entity: list[str],requested_config: CalculationTemplateRequest):    
        return  business_header_entity * len(requested_config.forecast)
    
    def apply_excel_month_forecast_headers(self, business_header_entity: list[str], requested_config: CalculationTemplateRequest,worksheet: openpyxl.worksheet.worksheet.Worksheet,forecast_months_headers: list[str]):
        total_column_need = (len(requested_config.forecast) * len(business_header_entity)) + 1
        for i in range(2, total_column_need, len(business_header_entity)):
            value = worksheet.cell(row=1, column=i, value=forecast_months_headers[int((i - 1) / len(business_header_entity))])
            value.alignment = Alignment(horizontal="center")
            worksheet.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i + len(business_header_entity) - 1)
            
    def apply_excel_business_headers(self, business_headers: list[str],worksheet: openpyxl.worksheet.worksheet.Worksheet):
        for i in range(len(business_headers)):
            worksheet.cell(row=2, column=2 + i, value=business_headers[i])
    
    def apply_excel_model_name(self, models: List[Model],worksheet: openpyxl.worksheet.worksheet.Worksheet):
        for i in range(len(models)):
            worksheet.cell(row=3 + i, column=1, value=models[i].name)
    
    def generate_calculation_excel_template(self, request: Request, query_params: CalculationTemplateRequest):
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet.cell(row=1, column=1, value="Slot Allocation")
        worksheet.cell(row=2, column=1, value="Model Name")
        
        business_header_entity = [
            "Take Off 100%",
            "BO",
            "Remaining",
            "Stock Pilot",
            "SOA",
            "OC",
            "Booking Prospect",
            "Forecast Order"
        ]
        
        forecast_months_header = self.generate_excel_month_header(query_params)
        business_headers = self.generate_excel_business_header(business_header_entity, query_params)
        
        self.apply_excel_month_forecast_headers(business_header_entity, query_params,worksheet, forecast_months_header)
        
        self.apply_excel_business_headers(business_headers,worksheet)
        
        models = self.master_repo.get_model_list(request)
        self.apply_excel_model_name(models,worksheet)
        
        # TODO need guidance with this
        temp_path = "/Users/apple/Documents/work/trilogic/code-project/his_vehicle_allocation_3a/va-backend/src/temp/calculation_template.xlsx"
        workbook.save(temp_path)
        
        return temp_path